import openpyxl
from openpyxl.styles import PatternFill


def GetRowCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_row)


def GetColumnCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_column)


def ReadData(file, SheetName, rowno, colmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(rowno, colmnno).value


def WriteData(file, SheetName, rowno, colmnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(rowno, colmnno).value = data
    workbook.save(file)


def FillGreenColor(file, SheetName, rowno, colmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    greenfill = PatternFill(start_color="60b212",
                            end_color="60b212",
                            fill_type="solid")
    sheet.cell(rowno, colmnno).fill = greenfill
    workbook.save(file)


def FillRedColor(file, SheetName, rowno, colmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    Redfill = PatternFill(start_color="ff0000",
                          end_color="ff0000",
                          fill_type="solid")
    sheet.cell(rowno, colmnno).fill = Redfill
    workbook.save(file)
